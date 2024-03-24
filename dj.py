#!/usr/bin/python3
# -*- coding: UTF-8 -*-


#一个非常麻烦的问题是，如果我们采取另建sheet的方式来录入生成的leg，
#那么读取active sheet页时遇到任何隐藏的问题，就难以发现了。
#用户只会一键复制走新sheet页的leg，贴到航网中完事。
#这种实现方式虽然最方便，但是并不安全。

import os
from openpyxl import Workbook
from openpyxl import load_workbook
import re
import datetime  #新增sheet页可能要增加时间戳，防止用户重复操作增加多个同名sheet页

standard_owner={'西北':'SIA','四川':'CTU','山东':'TAO','广东':'CAN',
'甘肃':'LHW','浙江':'NGB','江西':'KHN','安徽':'HFE','厦门':'XMN',
'山西':'TYN','北京':'PEK','江苏':'NKG','云南':'KMG','联航':'KN','一二三':'OTT',
'武汉':'WUH','海南':'SYX','上航':'FM','基地':'MU','西安':'SIA','中心':'FM'}

standard_city_code3={'如皋':'RUG','阆中':'LZG','乌拉特中旗':'WZQ','月牙湖':'YEH','兰坪白族普米族自治县':'LFH',
'东阳':'HEW','建德':'JDE','湘西':'DXJ','韶关':'HSC','鄂州':'EHU','郴州':'HCZ','成都天府':'TFU','于田':'YTW',
'菏泽':'HZA','武隆':'CQW','芜湖':'WHA','甘孜':'GZG','巫山':'WSK','莎车':'QSZ','巴中':'BZX','鄂尔多斯':'DSN',
'陇南':'LNL','遵义茅台':'WMT','信阳':'XAI','祁连':'HBQ','玉树':'YUS','腾冲':'TCZ','苏州':'SZV',
'拉斯卡尼亚斯':'LCS','自贡':'ZKL','昭苏':'ZFL','岳阳':'YYA','呼伦贝尔':'DWS','卡宾布里克':'CBZ',
'邢台':'XNT','布尔津':'KJI','且末':'IQM','大足':'DZU','蓬莱':'PNJ','安康':'AKA','朔州':'SZH',
'兴城':'XEN','长海':'CNI','新源':'NLT','兴宁':'XIN','二连浩特':'ERL','长沙':'CSX','怀化':'HJJ',
'蚌埠':'BFU','鄯善':'SXJ','富蕴':'FYN','老河口':'LHK','安阳':'AYN','梧州':'WUZ','依兰':'YLN',
'河池':'HCJ','乌兰察布':'UCB','三明':'SQJ','石河子':'SHF','博乐':'BPL','抚远':'FYJ',
'阿拉善额吉纳旗':'EJN','阿拉善右旗':'RHT','阿拉善左旗':'AXF','夏河':'GXH','邵阳':'WGN',
'梁平':'LIA','天水':'THQ','元谋':'YUA','广汉':'GHN','澜沧':'JMJ','白城':'DBC','承德':'CDE',
'琼海':'BAR','沧源':'CWJ','阿尔山':'YIE','五台山':'WUT','临汾':'LFQ','梅州':'MXZ',
'阿坝红原':'AHJ','黔江':'JIQ','日照':'RIZ','花土沟':'HTT','通化':'TNH','黑河':'HEK',
'神农架':'HPG','东营':'DOY','稻城':'DCY','锡林浩特':'XIL','武汉':'WUH','齐齐哈尔':'NDG',
'伊春':'LDS','吉林':'JIL','邯郸':'HDG','湛江':'ZHA','昭通':'ZAT','张掖':'YZY','盐城':'YNZ',
'营口':'YKH','宁蒗':'NLH','白山':'NBS','十堰':'WDS','日喀则':'RKZ','秦皇岛':'BPE','荔波':'LLB',
'哈密':'HMI','张家口':'ZQZ','遵义':'ZYI','池州':'JUH','衡阳':'HNY','汉中':'HZG','昌都':'BPX',
'六盘水':'LPF','凯里':'KJH','广元':'GYS','黎平':'HZH','上饶':'SQD','巴彦淖尔':'RLK',
'德令哈':'HXD','吕梁':'LLV','宜春':'YIC','果洛':'GMQ','惠州':'HUZ','昆明':'KMG','毕节':'BFJ',
'海拉尔':'HLD','洛阳':'LYA','柳州':'LZH','固原':'GYU','迪庆':'DIG','兰州':'LHW','乌鲁木齐':'URC',
'牡丹江':'MDG','和田':'HTN','康定':'KGT','阿勒泰':'AAT','延吉':'YNJ','潍坊':'WEF','常州':'CZX',
'长治':'CIH','百色':'AEB','珠海':'ZUH','金昌':'JIC','大理':'DLU','沈阳':'SHE','达州':'DZH',
'赣州':'KOW','包头':'BAV','吐鲁番':'TLQ','银川':'INC','南阳':'NNY','宜昌':'YIH','南昌':'KHN',
'广州':'CAN','佳木斯':'JMU','太原':'TYN','大连':'DLC','格尔木':'GOQ','通辽':'TGO','临沧':'LNJ',
'九寨沟':'JZH','海口':'HAK','呼和浩特':'HET','杭州':'HGH','南京':'NKG','阜阳':'FUG','赤峰':'CIF',
'兴义':'ACX','西双版纳':'JHG','榆林':'UYN','阿克苏':'AKU','延安':'ENY','南充':'NAO','南通':'NTG',
'加格达奇':'JGD','鞍山':'AOG','敦煌':'DNH','成都双流':'CTU','井冈山':'JGS','北京大兴':'PKX',
'大庆':'DQA','天津':'TSN','哈尔滨':'HRB','万州':'WXN','拉萨':'LXA','郑州':'CGO','保山':'BSD',
'长春':'CGQ','大同':'DAT','淮安':'HIA','北海':'BHY','阿里':'NGQ','唐山':'TVS','汕头':'SWA',
'乌兰浩特':'HLH','漠河':'OHE','徐州':'XUZ','舟山':'HSN','南宁':'NNG','泉州':'JJN','石家庄':'SJW',
'福州':'FOC','安庆':'AQG','扬州':'YTY','常德':'CGD','临沂':'LYI','合肥':'HFE','济南':'TNA',
'景德镇':'JDZ','伊宁':'YIN','厦门':'XMN','西安':'XIY','西昌':'XIC','襄阳':'XFN','无锡':'WUX',
'文山':'WNH','威海':'WEH','黄山':'TXN','铜仁':'TEN','青岛':'TAO','深圳':'SZX','普洱':'SYM',
'荆州':'SHS','上海虹桥':'SHA','北京首都':'PEK','北京南苑':'NAY','林芝':'LZY','芒市':'LUM',
'克拉玛依':'KRY','库尔勒':'KRL','安顺':'AVA','朝阳':'CHG','永州':'LLF','烟台':'YNT',
'喀什':'KHG','温州':'WNZ','庆阳':'IQN','库车':'KCA','丹东':'DDG','运城':'YCU',
'西宁':'XNN','锦州':'JNZ','三亚':'SYX','丽江':'LJG','绵阳':'MIG','连城':'LCX','佛山':'FUO',
'台州':'HYN','武夷山':'WUS','张家界':'DYG','泸州':'LZO','塔城':'TCG','衢州':'JUZ','鸡西':'JXA',
'上海浦东':'PVG','乌海':'WUA','重庆':'CKG','中卫':'ZHY','宜宾':'YBP','满洲里':'NZH','宁波':'NGB',
'桂林':'KWL','九江':'JIU','连云港':'LYG','贵阳':'KWE','嘉峪关':'JGN','攀枝花':'PZI','恩施':'ENH',
'义乌':'YIW','济宁':'JNG','松原':'YSQ','扎兰屯':'NZL','霍林郭勒':'HUO','若羌':'RQA',
'五大连池':'DTU','建三江':'JSJ','图木舒克':'TWC','青岛/胶东机场':'TAO','浦东':'PVG','虹桥':'SHA',
'大兴':'PKX','首都':'PEK','天府':'TFU','双流':'CTU'}

#关键词是否出现在某个tuple中，注意，这里是非常模糊的存在。比如'浦东'存在于(,,,"  上海浦东 ",,)也是可以的。
#这里采用正则表达式re search来解决模糊存在的问题
def keyword_in(keyword,tp):
	p=re.compile(keyword)
	#注意element里可能有数字，有日期格式等等，需要转换成字符串格式。不然会报错。
	for k,element in enumerate(tp):
		if p.search(str(element))!=None:
			return (k,element)
	return None

def std_owner(zhiguan):
	for nm in standard_owner:
		if nm in zhiguan:
			return standard_owner[nm]

def std_str_date(d):
	year=str(d.year)
	month=str(d.month)
	if len(month)==1:month='0'+month
	day=str(d.day)
	if len(day)==1:day='0'+day
	return year+'-'+month+'-'+day


class Standard_leg():
	def __init__(self,flight_num,start_date,end_date,takeoff,land,owner,plane_type):
		self.flight_num=flight_num
		self.start_date=start_date
		self.end_date=end_date
		self.freq='D'
		self.takeoff=takeoff
		self.land=land
		self.takeoff_time='0100'
		self.land_time='2300'
		self.owner=owner
		self.plane_type=str(plane_type)
		self.others=['L','D','Z']


if __name__=="__main__":

	excel_files=[]
	for root, dirs, files in os.walk("."):
		for name in files:
			if name[-5:]=='.xlsx' or name[-4:]=='.xls':
				excel_files.append(name)
#	print(excel_files)
#['4月计划（西北、本部、北京、四川、甘肃、厦门）-报市场.xlsx']

	for excel_file in excel_files:
		wb=load_workbook(excel_file)
		ws=wb.active

		correct_file_flag= False  #当找到某行含有"机号"关键词的时候，说明这是一份定检相关的正确文件，而不是什么乱七八糟的excel文件

		for k,row in enumerate(ws.iter_rows(values_only=True)):
			if keyword_in('机号',row)!=None:
				correct_file_flag=True  #这是一份正确的技术公司发来的定检文件，下面开始记录机号、机型、日期、执管等存在的列
				flight_num_index=keyword_in('机号',row)[0]
				plane_type_index=keyword_in('机型',row)[0]
				try:
					owner_index=keyword_in('执管',row)[0]
				except:
					owner_index=keyword_in('维护工厂',row)[0]
#				else:
#					print('请在表头维护执管列！')
				start_date_index=keyword_in('日期',row)[0]
				end_date_index=start_date_index+1
				takeoff_index=keyword_in('地点',row)[0]
				continue
			if correct_file_flag==True:
				if keyword_in('公司',row)!=None or keyword_in('基地',row)!=None or keyword_in('中心',row)!=None:#有定检信息的行必然含有"公司"或"基地"字样
					N=len(row)#从第N+1列开始填写航网格式的定检条
					flight_num=row[flight_num_index]
					days_cost=row[end_date_index]-row[start_date_index]
					if days_cost<=datetime.timedelta(5):
						flight_num='BA'+flight_num[-4:]
					else:
						flight_num='BC'+flight_num[-4:]
					start_date=row[start_date_index]
					end_date=row[end_date_index]
					takeoff=row[takeoff_index]
					land=row[takeoff_index]
					plane_type=row[plane_type_index]
					owner=row[owner_index]
					dj=Standard_leg(flight_num,start_date,end_date,standard_city_code3[takeoff],standard_city_code3[land],std_owner(owner),plane_type)
		#			print(dj.flight_num,dj.start_date,dj.end_date,dj.takeoff,dj.land,dj.owner,dj.plane_type)

					ws.cell(row=k+1,column=N+1).value=dj.flight_num
					ws.cell(row=k+1,column=N+3).value=std_str_date(dj.start_date)
					ws.cell(row=k+1,column=N+4).value=std_str_date(dj.end_date)
					ws.cell(row=k+1,column=N+5).value=dj.freq
					ws.cell(row=k+1,column=N+6).value=dj.takeoff
					ws.cell(row=k+1,column=N+7).value=dj.takeoff_time
					ws.cell(row=k+1,column=N+8).value=dj.land
					ws.cell(row=k+1,column=N+9).value=dj.land_time
					ws.cell(row=k+1,column=N+10).value=dj.owner
					ws.cell(row=k+1,column=N+11).value=dj.plane_type
					ws.cell(row=k+1,column=N+13).value=dj.others[0]
					ws.cell(row=k+1,column=N+14).value=dj.others[1]
					ws.cell(row=k+1,column=N+15).value=dj.others[2]

		wb.save(excel_file)  #原处覆盖并保存



